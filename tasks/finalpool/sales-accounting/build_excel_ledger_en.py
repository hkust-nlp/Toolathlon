#!/usr/bin/env python3
"""
Script to build complex ledger for digital products store in English
Uses openpyxl to directly create Excel files with over 100 transaction records
"""

import random
import datetime
from datetime import timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def generate_complex_ledger_data():
    """Generate complex digital products store ledger data in English"""
    
    # Product information
    products = {
        'iPhone 15': {'price': 7999, 'cost': 6500, 'category': 'Phone'},
        'iPhone 15 Pro': {'price': 9999, 'cost': 8000, 'category': 'Phone'},
        'iPhone 14': {'price': 6999, 'cost': 5500, 'category': 'Phone'},
        'MacBook Air': {'price': 8999, 'cost': 7200, 'category': 'Laptop'},
        'MacBook Pro': {'price': 12999, 'cost': 10500, 'category': 'Laptop'},
        'iPad': {'price': 4999, 'cost': 4000, 'category': 'Tablet'},
        'iPad Pro': {'price': 7999, 'cost': 6500, 'category': 'Tablet'},
        'AirPods Pro': {'price': 1899, 'cost': 1400, 'category': 'Accessories'},
        'AirPods': {'price': 1299, 'cost': 1000, 'category': 'Accessories'},
        'Apple Watch': {'price': 2999, 'cost': 2400, 'category': 'Smart Watch'},
        'iMac': {'price': 14999, 'cost': 12000, 'category': 'Desktop'},
        'Mac mini': {'price': 4999, 'cost': 4000, 'category': 'Desktop'},
        'Magic Keyboard': {'price': 999, 'cost': 700, 'category': 'Accessories'},
        'Magic Mouse': {'price': 699, 'cost': 500, 'category': 'Accessories'},
        'iPhone Charger': {'price': 299, 'cost': 200, 'category': 'Accessories'},
        'iPhone Case': {'price': 199, 'cost': 100, 'category': 'Accessories'},
        'MacBook Bag': {'price': 499, 'cost': 300, 'category': 'Accessories'},
        'iPad Keyboard': {'price': 1599, 'cost': 1200, 'category': 'Accessories'}
    }
    
    # Customer list
    customers = [
        'Zhang San', 'Li Si', 'Wang Wu', 'Zhao Liu', 'Liu Qi', 'Chen Ba', 'Yang Jiu', 'Huang Shi',
        'Zhou Yi', 'Wu Er', 'Zheng San', 'Sun Si', 'Zhu Wu', 'Hu Liu', 'Lin Qi', 'He Ba',
        'Gao Jiu', 'Liang Shi', 'Luo Shiyi', 'Song Shier', 'Xie Shisan', 'Tang Shisi', 'Han Shiwu', 'Feng Shiliu',
        'Yu Shiqi', 'Dong Shiba', 'Xiao Shijiu', 'Cheng Ershi', 'Cao Eryi', 'Yuan Erer', 'Deng Ersan', 'Xu Ersi',
        'Fu Erwu', 'Shen Erliu', 'Zeng Erqi', 'Peng Erba', 'Lu Erjiu', 'Su Sanshi', 'Lu Sanyi', 'Jiang Saner'
    ]
    
    # Suppliers
    suppliers = ['Supplier A', 'Supplier B', 'Supplier C', 'Supplier D', 'Supplier E']
    
    # Generate ledger data
    ledger_data = []
    
    # Generate historical data (December 1, 2023 to January 6, 2024)
    start_date = datetime.date(2023, 12, 1)
    current_date = datetime.date(2024, 1, 6)
    
    current = start_date
    transaction_count = 0
    
    # Set random seed for reproducible results
    random.seed(42)
    
    while current <= current_date and transaction_count < 120:
        # Skip some dates (simulate rest days)
        if random.random() < 0.1:  # 10% chance to skip this day
            current += timedelta(days=1)
            continue
            
        # Generate 2-5 transactions per day
        daily_transactions = random.randint(2, 5)
        
        for _ in range(daily_transactions):
            if transaction_count >= 120:
                break
                
            transaction_type = random.choices(['Sales', 'Purchase'], weights=[0.7, 0.3])[0]
            product = random.choice(list(products.keys()))
            quantity = random.randint(1, 3) if transaction_type == 'Sales' else random.randint(5, 20)
            
            if transaction_type == 'Sales':
                unit_price = products[product]['price']
                customer = random.choice(customers)
                # Random discounts
                if random.random() < 0.2:  # 20% chance for discount
                    discount = random.uniform(0.05, 0.15)
                    unit_price = int(unit_price * (1 - discount))
                    note = f"Discount {discount*100:.1f}%"
                else:
                    note = ""
                partner = customer
            else:  # Purchase
                unit_price = products[product]['cost']
                supplier = random.choice(suppliers)
                note = ""
                partner = supplier
            
            total_amount = unit_price * quantity
            
            ledger_data.append({
                'date': current.strftime('%Y-%m-%d'),
                'type': transaction_type,
                'product': product,
                'quantity': quantity,
                'unit_price': unit_price,
                'total': total_amount,
                'partner': partner,
                'note': note
            })
            
            transaction_count += 1
        
        current += timedelta(days=1)
    
    return ledger_data

def create_initial_ledger_excel():
    """Create initial ledger Excel file (historical data up to 2024-01-06) in English"""
    
    # Generate data
    ledger_data = generate_complex_ledger_data()
    print(f"Generated {len(ledger_data)} historical transaction records")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    
    # Set headers in English
    headers = ['Date', 'Type', 'Product', 'Quantity', 'Unit Price', 'Total', 'Customer/Supplier', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E1E1E1", end_color="E1E1E1", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_idx, record in enumerate(ledger_data, 2):
        ws.cell(row=row_idx, column=1, value=record['date'])
        ws.cell(row=row_idx, column=2, value=record['type'])
        ws.cell(row=row_idx, column=3, value=record['product'])
        ws.cell(row=row_idx, column=4, value=record['quantity'])
        ws.cell(row=row_idx, column=5, value=record['unit_price'])
        ws.cell(row=row_idx, column=6, value=record['total'])
        ws.cell(row=row_idx, column=7, value=record['partner'])
        ws.cell(row=row_idx, column=8, value=record['note'])
    
    # Adjust column widths
    column_widths = [12, 10, 15, 8, 10, 12, 15, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Save file
    file_path = "initial_workspace_en/Account_Book.xlsx"
    wb.save(file_path)
    print(f"Initial ledger saved: {file_path}")
    
    return ledger_data

def create_complete_ledger_excel(historical_data):
    """Create complete ledger Excel file (historical data + last week additions) in English"""
    
    # Last week's new data (January 7-10)
    last_week_data = [
        {
            'date': '2024-01-07',
            'type': 'Purchase',
            'product': 'iPhone 15 Pro',
            'quantity': 5,
            'unit_price': 8000,
            'total': 40000,
            'partner': 'Supplier A',
            'note': ''
        },
        {
            'date': '2024-01-08',
            'type': 'Purchase',
            'product': 'MacBook Air',
            'quantity': 3,
            'unit_price': 7200,
            'total': 21600,
            'partner': 'Supplier B',
            'note': ''
        },
        {
            'date': '2024-01-08',
            'type': 'Sales',
            'product': 'iPhone 15 Pro',
            'quantity': 2,
            'unit_price': 9999,
            'total': 19998,
            'partner': 'Wang Wu',
            'note': 'VIP Customer'
        },
        {
            'date': '2024-01-09',
            'type': 'Sales',
            'product': 'MacBook Air',
            'quantity': 1,
            'unit_price': 8999,
            'total': 8999,
            'partner': 'Zhao Liu',
            'note': 'New Customer'
        },
        {
            'date': '2024-01-10',
            'type': 'Sales',
            'product': 'AirPods Pro',
            'quantity': 3,
            'unit_price': 1899,
            'total': 5697,
            'partner': 'Liu Qi',
            'note': 'Returning Customer'
        }
    ]
    
    # Merge all data
    complete_data = historical_data + last_week_data
    print(f"Complete ledger contains {len(complete_data)} transaction records")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Complete Ledger"
    
    # Set headers in English
    headers = ['Date', 'Type', 'Product', 'Quantity', 'Unit Price', 'Total', 'Customer/Supplier', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E1E1E1", end_color="E1E1E1", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_idx, record in enumerate(complete_data, 2):
        ws.cell(row=row_idx, column=1, value=record['date'])
        ws.cell(row=row_idx, column=2, value=record['type'])
        ws.cell(row=row_idx, column=3, value=record['product'])
        ws.cell(row=row_idx, column=4, value=record['quantity'])
        ws.cell(row=row_idx, column=5, value=record['unit_price'])
        ws.cell(row=row_idx, column=6, value=record['total'])
        ws.cell(row=row_idx, column=7, value=record['partner'])
        ws.cell(row=row_idx, column=8, value=record['note'])
        
        # Highlight last week's new data
        if record['date'] >= '2024-01-07':
            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    # Adjust column widths
    column_widths = [12, 10, 15, 8, 10, 12, 15, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Save file
    file_path = "groundtruth_workspace_en/Account_Book_Complete.xlsx"
    wb.save(file_path)
    print(f"Complete ledger saved: {file_path}")
    
    return last_week_data

def generate_memory_json(last_week_data):
    """Generate memory.json knowledge graph data in English"""
    
    entities = []
    relations = []
    
    # Add product entities
    products_in_memory = ["iPhone 15 Pro", "MacBook Air", "AirPods Pro"]
    for product in products_in_memory:
        entities.append({
            "type": "entity",
            "name": product,
            "entityType": "Product",
            "observations": [
                f"Hot selling product in digital store",
                f"High-end product line",
                f"Key inventory management item"
            ]
        })
    
    # Add customer entities
    customers = ["Wang Wu", "Zhao Liu", "Liu Qi"]
    customer_details = {
        "Wang Wu": ["VIP customer", "Purchased two iPhone 15 Pro units", "Paid 19998 yuan in one transaction"],
        "Zhao Liu": ["New customer", "Purchased one MacBook Air unit", "Paid 8999 yuan"],
        "Liu Qi": ["Returning customer", "Purchased three AirPods Pro units", "Paid 5697 yuan"]
    }
    
    for customer in customers:
        entities.append({
            "type": "entity",
            "name": customer,
            "entityType": "Customer",
            "observations": customer_details[customer]
        })
    
    # Add supplier entities
    suppliers = ["Supplier A", "Supplier B"]
    supplier_details = {
        "Supplier A": ["Provides iPhone 15 Pro supply", "Purchase price 8000 yuan per unit", "3 years of cooperation"],
        "Supplier B": ["Provides MacBook Air supply", "Purchase price 7200 yuan per unit", "New business partner"]
    }
    
    for supplier in suppliers:
        entities.append({
            "type": "entity",
            "name": supplier,
            "entityType": "Supplier",
            "observations": supplier_details[supplier]
        })
    
    # Add transaction entities
    transaction_id = 1
    for trans in last_week_data:
        trans_name = f"Transaction{transaction_id:03d}"
        entities.append({
            "type": "entity",
            "name": trans_name,
            "entityType": f"{trans['type']} Transaction",
            "observations": [
                f"Date {trans['date']}",
                f"Product {trans['product']}",
                f"Quantity {trans['quantity']}",
                f"{'Customer' if trans['type'] == 'Sales' else 'Supplier'} {trans['partner']}",
                f"Amount {trans['total']} yuan"
            ]
        })
        
        # Add relationships
        relations.extend([
            {
                "type": "relation",
                "from": trans_name,
                "to": trans['partner'],
                "relationType": "Customer" if trans['type'] == 'Sales' else "Supplier"
            },
            {
                "type": "relation",
                "from": trans_name,
                "to": trans['product'],
                "relationType": "Product"
            }
        ])
        
        # Add customer/supplier to product relationships
        if trans['type'] == 'Sales':
            relations.append({
                "type": "relation",
                "from": trans['partner'],
                "to": trans['product'],
                "relationType": "Purchase"
            })
        else:
            relations.append({
                "type": "relation",
                "from": trans['partner'],
                "to": trans['product'],
                "relationType": "Supply"
            })
        
        transaction_id += 1
    
    memory_data = {
        "entities": entities,
        "relations": relations
    }
    
    # Save memory.json
    file_path = "initial_workspace_en/memory/memory.json"
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(memory_data, f, ensure_ascii=False, indent=2)
    
    print(f"Memory knowledge graph saved: {file_path}")
    print(f"Contains {len(entities)} entities and {len(relations)} relationships")
    
    return memory_data

def main():
    """Main function"""
    print("Starting to build complex digital products store ledger in English...")
    
    # Create initial ledger (historical data)
    historical_data = create_initial_ledger_excel()
    
    # Create complete ledger (historical + last week additions)
    last_week_data = create_complete_ledger_excel(historical_data)
    
    # Generate memory.json knowledge graph
    memory_data = generate_memory_json(last_week_data)
    
    print("\nLedger construction completed!")
    print(f"- Initial ledger: {len(historical_data)} historical records")
    print(f"- Complete ledger: {len(historical_data) + len(last_week_data)} records")
    print(f"- Knowledge graph: {len(memory_data['entities'])} entities, {len(memory_data['relations'])} relationships")

if __name__ == "__main__":
    main() 