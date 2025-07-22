#!/usr/bin/env python3
"""
构建数码产品店复杂账本的脚本
使用openpyxl直接创建Excel文件，包含超过100条交易记录
"""

import random
import datetime
from datetime import timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def generate_complex_ledger_data():
    """生成复杂的数码产品店账本数据"""
    
    # 商品信息
    products = {
        'iPhone 15': {'price': 7999, 'cost': 6500, 'category': '手机'},
        'iPhone 15 Pro': {'price': 9999, 'cost': 8000, 'category': '手机'},
        'iPhone 14': {'price': 6999, 'cost': 5500, 'category': '手机'},
        'MacBook Air': {'price': 8999, 'cost': 7200, 'category': '笔记本'},
        'MacBook Pro': {'price': 12999, 'cost': 10500, 'category': '笔记本'},
        'iPad': {'price': 4999, 'cost': 4000, 'category': '平板'},
        'iPad Pro': {'price': 7999, 'cost': 6500, 'category': '平板'},
        'AirPods Pro': {'price': 1899, 'cost': 1400, 'category': '配件'},
        'AirPods': {'price': 1299, 'cost': 1000, 'category': '配件'},
        'Apple Watch': {'price': 2999, 'cost': 2400, 'category': '智能手表'},
        'iMac': {'price': 14999, 'cost': 12000, 'category': '台式机'},
        'Mac mini': {'price': 4999, 'cost': 4000, 'category': '台式机'},
        'Magic Keyboard': {'price': 999, 'cost': 700, 'category': '配件'},
        'Magic Mouse': {'price': 699, 'cost': 500, 'category': '配件'},
        'iPhone 充电器': {'price': 299, 'cost': 200, 'category': '配件'},
        'iPhone 保护壳': {'price': 199, 'cost': 100, 'category': '配件'},
        'MacBook 包': {'price': 499, 'cost': 300, 'category': '配件'},
        'iPad 键盘': {'price': 1599, 'cost': 1200, 'category': '配件'}
    }
    
    # 客户名单
    customers = [
        '张三', '李四', '王五', '赵六', '刘七', '陈八', '杨九', '黄十',
        '周一', '吴二', '郑三', '孙四', '朱五', '胡六', '林七', '何八',
        '高九', '梁十', '罗十一', '宋十二', '谢十三', '唐十四', '韩十五', '冯十六',
        '于十七', '董十八', '萧十九', '程二十', '曹二一', '袁二二', '邓二三', '许二四',
        '傅二五', '沈二六', '曾二七', '彭二八', '吕二九', '苏三十', '卢三一', '蒋三二'
    ]
    
    # 供应商
    suppliers = ['供应商A', '供应商B', '供应商C', '供应商D', '供应商E']
    
    # 生成账本数据
    ledger_data = []
    
    # 生成历史数据（2023年12月1日到2024年1月6日）
    start_date = datetime.date(2023, 12, 1)
    current_date = datetime.date(2024, 1, 6)
    
    current = start_date
    transaction_count = 0
    
    # 设置随机种子以确保可重复的结果
    random.seed(42)
    
    while current <= current_date and transaction_count < 120:
        # 跳过一些日期（模拟休息日）
        if random.random() < 0.1:  # 10%概率跳过这一天
            current += timedelta(days=1)
            continue
            
        # 每天生成2-5笔交易
        daily_transactions = random.randint(2, 5)
        
        for _ in range(daily_transactions):
            if transaction_count >= 120:
                break
                
            transaction_type = random.choices(['销售', '进货'], weights=[0.7, 0.3])[0]
            product = random.choice(list(products.keys()))
            quantity = random.randint(1, 3) if transaction_type == '销售' else random.randint(5, 20)
            
            if transaction_type == '销售':
                unit_price = products[product]['price']
                customer = random.choice(customers)
                # 随机给予一些折扣
                if random.random() < 0.2:  # 20%概率有折扣
                    discount = random.uniform(0.05, 0.15)
                    unit_price = int(unit_price * (1 - discount))
                    note = f"折扣{discount*100:.1f}%"
                else:
                    note = ""
                partner = customer
            else:  # 进货
                unit_price = products[product]['cost']
                supplier = random.choice(suppliers)
                note = ""
                partner = supplier
            
            total_amount = unit_price * quantity
            
            ledger_data.append({
                'date': current.strftime('%Y-%m-%d'),
                'type': transaction_type,
                'product': product,
                'quantity': quantity,
                'unit_price': unit_price,
                'total': total_amount,
                'partner': partner,
                'note': note
            })
            
            transaction_count += 1
        
        current += timedelta(days=1)
    
    return ledger_data

def create_initial_ledger_excel():
    """创建初始账本Excel文件（历史数据到2024-01-06）"""
    
    # 生成数据
    ledger_data = generate_complex_ledger_data()
    print(f"生成了 {len(ledger_data)} 条历史交易记录")
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "账本"
    
    # 设置表头
    headers = ['日期', '交易类型', '商品名称', '数量', '单价', '总额', '客户/供应商', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E1E1E1", end_color="E1E1E1", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 写入数据
    for row_idx, record in enumerate(ledger_data, 2):
        ws.cell(row=row_idx, column=1, value=record['date'])
        ws.cell(row=row_idx, column=2, value=record['type'])
        ws.cell(row=row_idx, column=3, value=record['product'])
        ws.cell(row=row_idx, column=4, value=record['quantity'])
        ws.cell(row=row_idx, column=5, value=record['unit_price'])
        ws.cell(row=row_idx, column=6, value=record['total'])
        ws.cell(row=row_idx, column=7, value=record['partner'])
        ws.cell(row=row_idx, column=8, value=record['note'])
    
    # 调整列宽
    column_widths = [12, 10, 15, 8, 10, 12, 15, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # 保存文件
    file_path = "initial_workspace/账本.xlsx"
    wb.save(file_path)
    print(f"初始账本已保存: {file_path}")
    
    return ledger_data

def create_complete_ledger_excel(historical_data):
    """创建完整账本Excel文件（历史数据+上周新增）"""
    
    # 上周新增的数据（1月7-10日）
    last_week_data = [
        {
            'date': '2024-01-07',
            'type': '进货',
            'product': 'iPhone 15 Pro',
            'quantity': 5,
            'unit_price': 8000,
            'total': 40000,
            'partner': '供应商A',
            'note': ''
        },
        {
            'date': '2024-01-08',
            'type': '进货',
            'product': 'MacBook Air',
            'quantity': 3,
            'unit_price': 7200,
            'total': 21600,
            'partner': '供应商B',
            'note': ''
        },
        {
            'date': '2024-01-08',
            'type': '销售',
            'product': 'iPhone 15 Pro',
            'quantity': 2,
            'unit_price': 9999,
            'total': 19998,
            'partner': '王五',
            'note': 'VIP客户'
        },
        {
            'date': '2024-01-09',
            'type': '销售',
            'product': 'MacBook Air',
            'quantity': 1,
            'unit_price': 8999,
            'total': 8999,
            'partner': '赵六',
            'note': '新客户'
        },
        {
            'date': '2024-01-10',
            'type': '销售',
            'product': 'AirPods Pro',
            'quantity': 3,
            'unit_price': 1899,
            'total': 5697,
            'partner': '刘七',
            'note': '老客户'
        }
    ]
    
    # 合并所有数据
    complete_data = historical_data + last_week_data
    print(f"完整账本包含 {len(complete_data)} 条交易记录")
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "完整账本"
    
    # 设置表头
    headers = ['日期', '交易类型', '商品名称', '数量', '单价', '总额', '客户/供应商', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E1E1E1", end_color="E1E1E1", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 写入数据
    for row_idx, record in enumerate(complete_data, 2):
        ws.cell(row=row_idx, column=1, value=record['date'])
        ws.cell(row=row_idx, column=2, value=record['type'])
        ws.cell(row=row_idx, column=3, value=record['product'])
        ws.cell(row=row_idx, column=4, value=record['quantity'])
        ws.cell(row=row_idx, column=5, value=record['unit_price'])
        ws.cell(row=row_idx, column=6, value=record['total'])
        ws.cell(row=row_idx, column=7, value=record['partner'])
        ws.cell(row=row_idx, column=8, value=record['note'])
        
        # 高亮显示上周新增的数据
        if record['date'] >= '2024-01-07':
            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    # 调整列宽
    column_widths = [12, 10, 15, 8, 10, 12, 15, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # 保存文件
    file_path = "groundtruth_workspace/账本_完整.xlsx"
    wb.save(file_path)
    print(f"完整账本已保存: {file_path}")
    
    return last_week_data

def generate_memory_json(last_week_data):
    """生成memory.json知识图谱数据"""
    
    entities = []
    relations = []
    
    # 添加商品实体
    products_in_memory = ["iPhone 15 Pro", "MacBook Air", "AirPods Pro"]
    for product in products_in_memory:
        entities.append({
            "type": "entity",
            "name": product,
            "entityType": "商品",
            "observations": [
                f"数码产品店热销商品",
                f"高端产品线",
                f"库存管理重点商品"
            ]
        })
    
    # 添加客户实体
    customers = ["王五", "赵六", "刘七"]
    customer_details = {
        "王五": ["VIP客户", "购买iPhone 15 Pro两台", "一次性支付19998元"],
        "赵六": ["新客户", "购买MacBook Air一台", "支付8999元"],
        "刘七": ["老客户", "购买AirPods Pro三台", "支付5697元"]
    }
    
    for customer in customers:
        entities.append({
            "type": "entity",
            "name": customer,
            "entityType": "客户",
            "observations": customer_details[customer]
        })
    
    # 添加供应商实体
    suppliers = ["供应商A", "供应商B"]
    supplier_details = {
        "供应商A": ["提供iPhone 15 Pro进货", "进货价格8000元/台", "合作3年"],
        "供应商B": ["提供MacBook Air进货", "进货价格7200元/台", "新合作伙伴"]
    }
    
    for supplier in suppliers:
        entities.append({
            "type": "entity",
            "name": supplier,
            "entityType": "供应商",
            "observations": supplier_details[supplier]
        })
    
    # 添加交易实体
    transaction_id = 1
    for trans in last_week_data:
        trans_name = f"交易{transaction_id:03d}"
        entities.append({
            "type": "entity",
            "name": trans_name,
            "entityType": f"{trans['type']}交易",
            "observations": [
                f"日期{trans['date']}",
                f"商品{trans['product']}",
                f"数量{trans['quantity']}",
                f"{'客户' if trans['type'] == '销售' else '供应商'}{trans['partner']}",
                f"金额{trans['total']}元"
            ]
        })
        
        # 添加关系
        relations.extend([
            {
                "type": "relation",
                "from": trans_name,
                "to": trans['partner'],
                "relationType": "客户" if trans['type'] == '销售' else "供应商"
            },
            {
                "type": "relation",
                "from": trans_name,
                "to": trans['product'],
                "relationType": "商品"
            }
        ])
        
        # 添加客户/供应商与商品的关系
        if trans['type'] == '销售':
            relations.append({
                "type": "relation",
                "from": trans['partner'],
                "to": trans['product'],
                "relationType": "购买"
            })
        else:
            relations.append({
                "type": "relation",
                "from": trans['partner'],
                "to": trans['product'],
                "relationType": "供应"
            })
        
        transaction_id += 1
    
    memory_data = {
        "entities": entities,
        "relations": relations
    }
    
    # 保存memory.json
    file_path = "initial_workspace/memory.json"
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(memory_data, f, ensure_ascii=False, indent=2)
    
    print(f"Memory知识图谱已保存: {file_path}")
    print(f"包含 {len(entities)} 个实体，{len(relations)} 个关系")
    
    return memory_data

def main():
    """主函数"""
    print("开始构建数码产品店复杂账本...")
    
    # 创建初始账本（历史数据）
    historical_data = create_initial_ledger_excel()
    
    # 创建完整账本（历史+上周新增）
    last_week_data = create_complete_ledger_excel(historical_data)
    
    # 生成memory.json知识图谱
    memory_data = generate_memory_json(last_week_data)
    
    print("\n账本构建完成！")
    print(f"- 初始账本: {len(historical_data)} 条历史记录")
    print(f"- 完整账本: {len(historical_data) + len(last_week_data)} 条记录")
    print(f"- 知识图谱: {len(memory_data['entities'])} 个实体，{len(memory_data['relations'])} 个关系")

if __name__ == "__main__":
    main() 