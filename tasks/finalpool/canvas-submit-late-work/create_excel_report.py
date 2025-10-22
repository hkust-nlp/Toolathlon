#!/usr/bin/env python3
"""
Cinema Culture Appreciation Course - Grade Summary Generator
This script creates a beautiful Excel spreadsheet with student grades and statistics.
"""

import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

def create_grade_summary_excel():
    """Create a comprehensive Excel file with student grades and statistics"""
    
    # Load user configuration data
    with open('/ssddata/xiaochen/workspace/toolathlon/tasks/xiaochen/canvas_collect_work_data/files/user_config.json', 'r') as f:
        data = json.load(f)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade Summary"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    student_font = Font(name='Arial', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Score fills for different grade ranges
    excellent_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')  # Green
    good_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')      # Orange
    needs_improve_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
    missing_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Gray
    
    # Course title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Cinema Culture Appreciation (FILM101) - Grade Summary'
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # Date generated
    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(name='Arial', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Student ID', 'Student Name', 'Assignment 1', 'Assignment 2', 'Assignment 3', 
              'Assignment 4', 'Assignment 5', 'Assignment 6', 'Average', 'Completion']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Assignment names as second header row
    assignment_names = [
        '', '',  # Student ID and Name columns
        'Classic Film Analysis',
        'Film History & Evolution', 
        'Genre Analysis',
        'Cinematography & Visual',
        'International Cinema',
        'Contemporary Trends',
        '', ''  # Average and Completion columns
    ]
    
    for col, name in enumerate(assignment_names, 1):
        if name:  # Only for assignment columns
            cell = ws.cell(row=5, column=col)
            cell.value = name
            cell.font = Font(name='Arial', size=8, italic=True)
            cell.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Student data
    for row, student in enumerate(data['students'], 6):
        # Student ID
        ws.cell(row=row, column=1, value=student['id']).border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        # Student Name
        ws.cell(row=row, column=2, value=student['name']).border = border
        ws.cell(row=row, column=2).font = Font(name='Arial', size=10, bold=True)
        
        # Assignment scores
        scores = list(student['assignment_scores'].values())
        for col, score in enumerate(scores, 3):
            cell = ws.cell(row=row, column=col, value=score if score > 0 else 'Not Submitted')
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            # Color coding based on score
            if score == 0:
                cell.fill = missing_fill
                cell.font = Font(name='Arial', size=10, color='666666')
            elif score >= 90:
                cell.fill = excellent_fill
                cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            elif score >= 75:
                cell.fill = good_fill
                cell.font = Font(name='Arial', size=10, bold=True)
            else:
                cell.fill = needs_improve_fill
                cell.font = Font(name='Arial', size=10, color='FFFFFF')
        
        # Average score
        avg_cell = ws.cell(row=row, column=9, value=f"{student['average_score']:.1f}")
        avg_cell.border = border
        avg_cell.alignment = Alignment(horizontal='center')
        avg_cell.font = Font(name='Arial', size=10, bold=True)
        
        if student['average_score'] >= 80:
            avg_cell.fill = excellent_fill
            avg_cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        elif student['average_score'] >= 60:
            avg_cell.fill = good_fill
        else:
            avg_cell.fill = needs_improve_fill
            avg_cell.font = Font(name='Arial', size=10, color='FFFFFF')
        
        # Completion rate
        completion_cell = ws.cell(row=row, column=10, value=student['completion_rate'])
        completion_cell.border = border
        completion_cell.alignment = Alignment(horizontal='center')
        completion_cell.font = Font(name='Arial', size=10, bold=True)
    
    # Statistics section
    stats_row = len(data['students']) + 8
    
    ws.merge_cells(f'A{stats_row}:J{stats_row}')
    ws[f'A{stats_row}'] = 'Course Statistics'
    ws[f'A{stats_row}'].font = Font(name='Arial', size=14, bold=True)
    ws[f'A{stats_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{stats_row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws[f'A{stats_row}'].alignment = Alignment(horizontal='center')
    
    stats_data = [
        ['Total Students', data['statistics']['total_students']],
        ['Class Average', f"{data['statistics']['class_average']:.1f}%"],
        ['Students with Complete Assignments', len(data['statistics']['fully_completed_students'])],
        ['Students with Incomplete Assignments', len(data['statistics']['students_with_incomplete_assignments'])],
        ['Completion Rate', f"{len(data['statistics']['fully_completed_students'])/data['statistics']['total_students']*100:.1f}%"]
    ]
    
    for i, (label, value) in enumerate(stats_data, stats_row + 1):
        ws.cell(row=i, column=1, value=label).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=i, column=2, value=value).font = Font(name='Arial', size=10)
    
    # Legend
    legend_row = stats_row + 7
    ws.merge_cells(f'A{legend_row}:J{legend_row}')
    ws[f'A{legend_row}'] = 'Grade Legend'
    ws[f'A{legend_row}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'A{legend_row}'].alignment = Alignment(horizontal='center')
    
    legend_items = [
        ('Excellent (90-100)', excellent_fill, 'FFFFFF'),
        ('Good (75-89)', good_fill, '000000'),
        ('Needs Improvement (0-74)', needs_improve_fill, 'FFFFFF'),
        ('Not Submitted', missing_fill, '666666')
    ]
    
    for i, (label, fill, font_color) in enumerate(legend_items, legend_row + 1):
        cell = ws.cell(row=i, column=1, value=label)
        cell.fill = fill
        cell.font = Font(name='Arial', size=10, bold=True, color=font_color)
        cell.border = border
    
    # Adjust column widths
    column_widths = [12, 20, 15, 18, 15, 18, 16, 16, 10, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Create second sheet for incomplete assignments
    ws2 = wb.create_sheet("Incomplete Assignments")
    
    # Header for incomplete assignments sheet
    ws2.merge_cells('A1:D1')
    ws2['A1'] = 'Students with Incomplete Assignments'
    ws2['A1'].font = Font(name='Arial', size=14, bold=True)
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2['A1'].fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    ws2['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    
    incomplete_headers = ['Student Name', 'Email', 'Missing Assignments', 'Completion Rate']
    for col, header in enumerate(incomplete_headers, 1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add students with incomplete assignments
    row = 4
    for student in data['students']:
        if student['email'] in data['statistics']['students_with_incomplete_assignments']:
            missing_assignments = [name for name, score in student['assignment_scores'].items() if score == 0]
            
            ws2.cell(row=row, column=1, value=student['name']).border = border
            ws2.cell(row=row, column=2, value=student['email']).border = border
            ws2.cell(row=row, column=3, value=', '.join([name.split(': ')[0] for name in missing_assignments])).border = border
            ws2.cell(row=row, column=4, value=student['completion_rate']).border = border
            
            # Color code based on completion rate
            completion_num = int(student['completion_rate'].split('/')[0])
            if completion_num <= 2:
                for col in range(1, 5):
                    ws2.cell(row=row, column=col).fill = needs_improve_fill
                    ws2.cell(row=row, column=col).font = Font(color='FFFFFF')
            elif completion_num <= 4:
                for col in range(1, 5):
                    ws2.cell(row=row, column=col).fill = good_fill
            
            row += 1
    
    # Adjust column widths for second sheet
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 15
    
    # Save the workbook
    output_path = '/ssddata/xiaochen/workspace/toolathlon/tasks/xiaochen/canvas_collect_work_data/initial_workspace/Cinema_Culture_Grade_Summary.xlsx'
    wb.save(output_path)
    
    print(f"✅ Grade summary Excel file created successfully: {output_path}")
    print(f"📊 Total students: {data['statistics']['total_students']}")
    print(f"📈 Class average: {data['statistics']['class_average']:.1f}%")
    print(f"✅ Students with complete assignments: {len(data['statistics']['fully_completed_students'])}")
    print(f"⚠️  Students with incomplete assignments: {len(data['statistics']['students_with_incomplete_assignments'])}")
    
    return output_path

if __name__ == "__main__":
    create_grade_summary_excel()