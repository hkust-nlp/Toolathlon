import os
import re
import openpyxl

# List all files in the directory
all_files = os.listdir('.')

# Filter files related to Operating Systems HW3
hw3_os_files = [f for f in all_files if re.search(r"OperatingSystems|OS|OS-", f) and re.search(r"3|Assignment3|HW3", f)]

# Read correspondence table
wb = openpyxl.load_workbook('CollegeNameID.xlsx')

# Combine all sheets for lookup
student_info = {}
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Only take first three columns
        if row and len(row) >= 3:
            name, student_id, college = row[:3]
            student_info[name] = (student_id, college)

# Function to determine language folder
def get_language_folder(filename):
    if filename.endswith('.c'):
        return 'C'
    elif filename.endswith('.rs'):
        return 'Rust'
    elif filename.endswith('.py'):
        return 'Python'
    else:
        return None

# Rename and move files
base_dir = '.'
new_base_dir = './os_hw3'
renamed_files = []
for old_name in hw3_os_files:
    # Extract student name from filename
    # Student name is assumed to be the first part before '_' or '-'
    if '_' in old_name:
        student_name = old_name.split('_')[0]
    else:
        student_name = old_name.split('-')[0]
    
    # Lookup student info
    if student_name in student_info:
        student_id, college = student_info[student_name]
        # Compose new name
        new_name = f"{student_name}-{college}-{student_id}-OS-HW3" + os.path.splitext(old_name)[1]
        # Determine language folder
        lang_folder = get_language_folder(old_name)
        if lang_folder:
            new_path = os.path.join(new_base_dir, lang_folder, new_name)
            old_path = os.path.join(base_dir, old_name)
            os.rename(old_path, new_path)
            renamed_files.append((old_name, new_name, lang_folder))

# Delete remaining original code files
all_files = os.listdir(base_dir)
for file in all_files:
    if file != 'CollegeNameID.xlsx' and not file.startswith('os_hw3') and not os.path.isdir(file):
        os.remove(file)

renamed_files
