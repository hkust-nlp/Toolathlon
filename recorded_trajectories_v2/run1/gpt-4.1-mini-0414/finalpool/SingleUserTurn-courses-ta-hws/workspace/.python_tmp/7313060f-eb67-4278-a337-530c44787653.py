import os
import re

# List of files to rename
hw3_os_files = ['David Miller_OperatingSystems_HW3.py', 'Robert Martinez-OS-Homework3.rs', 'Matthew Lee-OperatingSystems-HW3.rs', 'Samuel King_OperatingSystems_HW3.py', 'Ryan Scott_OperatingSystems_Assignment3.rs', 'Helen Edwards_OperatingSystems_HW3-VirtualMemory.py', 'Mark Cook-OS-Homework3.py', 'Frank Morgan-OperatingSystems-Lab3.rs', 'Larry Ward_OperatingSystems_HW3.rs', 'Louis Brooks_OS_Lab3-Threading.c', 'Phillip Sanders_OS_Lab3.py', 'Jeffrey Powell-OperatingSystems-Lab3.py', 'Nina Bryant-OS-Project3.rs', 'Howard Alexander-OperatingSystems-HW3-ProcessManagement.rs', 'Yvonne Wells_OperatingSystems_HW3.py', 'Michael Johnson-OS-Homework3-Threading.rs', 'James Brown-OperatingSystems-HW3.py', 'William Davis_OS_Assignment3.rs', 'Kevin Wilson_OperatingSystems_HW3.rs', 'Thomas Anderson_OperatingSystems_HW3.c', 'Andrew Taylor-OperatingSystems-Lab3.c', 'Daniel Thompson-OperatingSystems-HW3.c', 'John Garcia_OS_Assignment3.c', 'Christopher Rodriguez_OS_Lab3.py', 'Steven Lewis_OperatingSystems_HW3-Threading.py', 'Anthony Walker_OperatingSystems_HW3.py', 'Alexander Hall_OperatingSystems_Assignment3.rs', 'Nicholas Allen-OS-Homework3.py', 'Benjamin Young-OS-Assignment3.py', 'Joshua Wright_OS_Lab3-DeadlockDetection.rs', 'Eric Green-OS-Assignment3.c', 'Sophia Adams_OperatingSystems_Assignment3.c', 'Grace Baker-OS-Project3-VirtualMemory.rs', 'Nathan Gonzalez-OS-Assignment3.rs', 'Jason Nelson_OS_Project3.rs', 'Emily Carter_OS_Lab3.rs', 'Lucas Mitchell_OS_Assignment3.rs', 'Michelle Perez-OperatingSystems-Lab3.rs', 'Jennifer Roberts-OperatingSystems-HW3.rs', 'Alex Turner_OS_Project3.c', 'Crystal Phillips-OS-Homework3.c', 'Emma Campbell-OperatingSystems-Lab3.rs', 'Tony Parker_OperatingSystems_Assignment3.py', 'Tina Evans_OS_Assignment3.c', 'Charlie Collins-OperatingSystems-HW3.c', 'Jack Stewart-OS-Project3.rs', 'Victor Sanchez_OS_Lab3.py', 'Jacob Morris_OS_Lab3-MemoryManagement.rs', 'Penny Rogers_OperatingSystems_HW3.rs', 'Peter Reed_OperatingSystems_Assignment3.c', 'Leo Bell-OS-Homework3.c', 'Jerry Murphy-OS-Project3-Scheduling.py', 'Henry Bailey-OS-Assignment3.py', 'Celia Rivera_OperatingSystems_HW3.rs', 'Leonard Cooper-OS-Project3.py', 'Hannah Richardson-OS-Homework3.c', 'Yvonne Cox-OperatingSystems-HW3-Scheduling.rs', 'Linda Torres_OS_Assignment3.c', 'Paul Peterson-OS-Project3-VirtualMemory.c', 'Theodore Gray_OperatingSystems_Assignment3.py', 'Quinn Ramirez_OS_Assignment3-FileSystem.rs', 'Helen James-OS-Project3.py', 'Cassie Watson_OS_Lab3.c', 'Julian Kelly_OS_Lab3.c', 'Lillian Price-OperatingSystems-Lab3.c', 'Jake Bennett-OperatingSystems-Lab3.c', 'Hugh Wood-OS-Homework3.c', 'Hazel Barnes-OperatingSystems-Lab3.py', 'Felix Ross-OS-Homework3.rs', 'Tiffany Henderson_OS_Project3.py', 'Bobby Coleman_OS_Project3-Threading.c', 'Jesse Jenkins-OperatingSystems-Lab3.py', 'Logan Perry-OperatingSystems-HW3.rs', 'Yvette Long-OperatingSystems-HW3.rs', 'Lynn Patterson-OS-Homework3.py', 'Ray Hughes-OS-Homework3.py', 'Ying Flores-OperatingSystems-Lab3.c', 'Catherine Washington_OS_Assignment3-Scheduling.c', 'Lily Butler-OperatingSystems-HW3.rs', 'Jenna Simmons-OS-Assignment3.c', 'Ruby Foster_OperatingSystems_Assignment3.c', 'Quincy Gonzales-OS-Homework3.rs', 'Theresa Russell-OS-Homework3-ProcessManagement.c', 'Rose Griffin_OS_Project3-MemoryManagement.rs', 'Lila Diaz-OperatingSystems-HW3.py', 'Serena Hayes-OS-Assignment3.rs', 'Susan Myers-OS-Assignment3.py', 'Riley Ford-OperatingSystems-Lab3.py', 'Sarah Hamilton-OS-Homework3.py', 'Kenny Graham-OS-Project3.c', 'Keith Sullivan_OS_Project3.c', 'Patrick Wallace_OS_Project3.c', 'Lena Woods_OperatingSystems_Assignment3.py', 'Ulysses Cole_OS_Project3.c', 'Yolanda West_OS_Assignment3.c', 'Marlene Jordan-OS-Project3-MemoryManagement.rs', 'Yvette Stone-OperatingSystems-Lab3.py', 'Bob Freeman-OperatingSystems-Lab3.c', 'Boris Webb-OS-Homework3.rs', 'Cindy Tucker-OS-Assignment3.py']

# Read correspondence table
import openpyxl
wb = openpyxl.load_workbook('CollegeNameID.xlsx')

# Combine all sheets for lookup
student_info = {}
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, student_id, college = row
        student_info[name] = (student_id, college)

# Function to determine language folder
def get_language_folder(filename):
    if filename.endswith('.c'):
        return 'C'
    elif filename.endswith('.rs'):
        return 'Rust'
    elif filename.endswith('.py'):
        return 'Python'
    else:
        return None

# Rename and move files
base_dir = '.'
new_base_dir = './os_hw3'
renamed_files = []
for old_name in hw3_os_files:
    # Extract student name from filename
    # Student name is assumed to be the first part before '_' or '-'
    if '_' in old_name:
        student_name = old_name.split('_')[0]
    else:
        student_name = old_name.split('-')[0]
    
    # Lookup student info
    if student_name in student_info:
        student_id, college = student_info[student_name]
        # Compose new name
        new_name = f"{student_name}-{college}-{student_id}-OS-HW3" + os.path.splitext(old_name)[1]
        # Determine language folder
        lang_folder = get_language_folder(old_name)
        if lang_folder:
            new_path = os.path.join(new_base_dir, lang_folder, new_name)
            old_path = os.path.join(base_dir, old_name)
            os.rename(old_path, new_path)
            renamed_files.append((old_name, new_name, lang_folder))

# Delete remaining original code files
all_files = os.listdir(base_dir)
for file in all_files:
    if file != 'CollegeNameID.xlsx' and not file.startswith('os_hw3'):
        os.remove(file)

renamed_files
