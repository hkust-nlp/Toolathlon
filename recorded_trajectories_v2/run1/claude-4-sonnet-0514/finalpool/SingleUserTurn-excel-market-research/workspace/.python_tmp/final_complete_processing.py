import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Read the Market_Data.xlsx file
file_path = '/ssddata/mcpbench/wenshuo/eval/final_new/mcpbench_dev/recorded_trajectories_v2/run1/claude-4-sonnet-0514/finalpool/SingleUserTurn-excel-market-research/workspace/Market_Data.xlsx'

# Load the raw data sheet
raw_data_df = pd.read_excel(file_path, sheet_name='RawData', header=None)

print("Processing complete data conversion and growth rate calculation...")

# Get the years from raw data and calculate appliance values
years = []
appliance_values = []
electric_percentages = []
construction_percentages = []
furniture_percentages = []

# Extract years and calculate appliance values for each year
for row_idx in range(4, len(raw_data_df)):
    year = raw_data_df.iloc[row_idx, 0]
    if pd.notna(year) and isinstance(year, (int, float)):
        years.append(int(year))
        
        # Appliance: Electric * 0.5 + Furniture * 0.2 + Construction * 0.3
        electric_value = raw_data_df.iloc[row_idx, 7]  # Column 7 is Electric
        furniture_value = raw_data_df.iloc[row_idx, 10]  # Column 10 is Furniture  
        construction_value = raw_data_df.iloc[row_idx, 11]  # Column 11 is Construction
        
        # Handle missing values
        electric_value = 0 if pd.isna(electric_value) else electric_value
        furniture_value = 0 if pd.isna(furniture_value) else furniture_value
        construction_value = 0 if pd.isna(construction_value) else construction_value
        
        appliance_value = (electric_value * 0.5 + 
                          furniture_value * 0.2 + 
                          construction_value * 0.3)
        
        appliance_values.append(appliance_value)
        
        # Calculate percentages for the format file
        if appliance_value > 0:
            electric_pct = (electric_value * 0.5) / appliance_value * 100
            furniture_pct = (furniture_value * 0.2) / appliance_value * 100
            construction_pct = (construction_value * 0.3) / appliance_value * 100
        else:
            electric_pct = furniture_pct = construction_pct = 0
            
        electric_percentages.append(electric_pct)
        furniture_percentages.append(furniture_pct)
        construction_percentages.append(construction_pct)

# Calculate the annual growth rates for the Appliance category from 2015 to 2024
growth_rates = []

# Find the indices for years 2015-2024
start_idx = years.index(2015)
end_idx = years.index(2024)

for i in range(start_idx, end_idx):
    current_value = appliance_values[i]
    next_value = appliance_values[i + 1]
    
    if current_value != 0:
        growth_rate = ((next_value - current_value) / current_value) * 100
    else:
        growth_rate = 0
    
    growth_rates.append(growth_rate)

# Create the results DataFrame for the output file
results_data = []
for i in range(start_idx, end_idx + 1):
    year = years[i]
    electric_pct = electric_percentages[i]
    construction_pct = construction_percentages[i]
    furniture_pct = furniture_percentages[i]
    
    if i < end_idx:  # Growth rate for this year to next year
        growth_rate = growth_rates[i - start_idx]
    else:  # No growth rate for the last year
        growth_rate = None
    
    results_data.append([year, electric_pct, construction_pct, furniture_pct, growth_rate])

# Create DataFrame
results_df = pd.DataFrame(results_data, columns=['Year', 'Electric %', 'Construction %', 'Furniture %', 'Growth Rate %'])

# Create the Excel file with the results
output_file = '/ssddata/mcpbench/wenshuo/eval/final_new/mcpbench_dev/recorded_trajectories_v2/run1/claude-4-sonnet-0514/finalpool/SingleUserTurn-excel-market-research/workspace/growth_rate.xlsx'

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Growth Rate Analysis"

# Add headers
headers = ['Year', 'Electric %', 'Construction %', 'Furniture %', 'Growth Rate %']
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Add data
for row_idx, (_, row_data) in enumerate(results_df.iterrows(), 2):
    for col_idx, value in enumerate(row_data, 1):
        if pd.isna(value):
            ws.cell(row=row_idx, column=col_idx, value=None)
        else:
            ws.cell(row=row_idx, column=col_idx, value=value)

# Save the workbook
wb.save(output_file)
print(f"Excel file saved as: {output_file}")

# Verify the file can be loaded with load_workbook(file_path, data_only=True)
try:
    test_wb = load_workbook(output_file, data_only=True)
    print("✓ File can be loaded with load_workbook(file_path, data_only=True)")
    
    # Show the data
    test_ws = test_wb.active
    print("\nFile contents verification:")
    for row in test_ws.iter_rows(max_row=11, values_only=True):
        print(row)
        
    test_wb.close()
except Exception as e:
    print(f"✗ Error loading file: {e}")

print("\nSummary:")
print(f"- Processed Appliance category data from 2015 to 2024")
print(f"- Appliance category is calculated as: Electric * 0.5 + Furniture * 0.2 + Construction * 0.3")
print(f"- Annual growth rates calculated for each year from 2015 to 2024")
print(f"- Highest growth rate: {max([g for g in growth_rates if g is not None]):.2f}% (2019 to 2020)")
print(f"- Output file: {output_file}")