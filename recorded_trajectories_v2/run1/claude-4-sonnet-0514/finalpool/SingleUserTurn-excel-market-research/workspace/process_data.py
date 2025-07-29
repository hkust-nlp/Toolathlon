from openpyxl import load_workbook, Workbook

# Load the Market_Data.xlsx file
wb = load_workbook('/ssddata/mcpbench/zhaojian/mcpbench_dev/recorded_trajectories_v2/run1/claude-4-sonnet-0514/finalpool/SingleUserTurn-excel-market-research/workspace/Market_Data.xlsx', data_only=True)

# Read Methodology sheet
methodology_ws = wb['Methodology']
raw_data_ws = wb['RawData']

# Extract raw data
years = []
raw_data = []

# Read years and data from RawData sheet (rows 5-23, columns A-P)
for row in raw_data_ws.iter_rows(min_row=5, max_row=23, min_col=1, max_col=16):
    year = row[0].value
    if year is not None:
        years.append(year)
        row_data = []
        for cell in row[1:]:  # Skip year column
            value = cell.value if cell.value is not None else 0
            row_data.append(value)
        raw_data.append(row_data)

# Column indices in raw_data:
# 0: Agriculture, 1: Oil_Gas, 2: Food, 3: Metals, 4: Electronic, 5: Electric, 
# 6: Auto_Production, 7: Transportation, 8: Furniture, 9: Construction, 
# 10: Industry, 11: Steel_Production, 12: Air_Bus, 13: Boeing, 14: MRO

# Calculate Appliance category values based on the conversion weights
# From methodology: Appliance = Electric * 0.5 + Furniture * 0.2 + Construction * 0.3
appliance_values = []

for i, year in enumerate(years):
    electric_val = raw_data[i][5]  # Electric column (F in RawData)
    furniture_val = raw_data[i][8]  # Furniture column (J in RawData)
    construction_val = raw_data[i][9]  # Construction column (K in RawData)
    
    # Convert Construction from bn USD to mn USD (multiply by 1000)
    construction_val_mn = construction_val * 1000 if construction_val != 0 else 0
    
    appliance_val = electric_val * 0.5 + furniture_val * 0.2 + construction_val_mn * 0.3
    appliance_values.append(appliance_val)

# Calculate annual growth rates
growth_rates = []
for i in range(1, len(appliance_values)):
    if appliance_values[i-1] != 0:
        growth_rate = (appliance_values[i] - appliance_values[i-1]) / appliance_values[i-1] * 100
        growth_rates.append(growth_rate)
    else:
        growth_rates.append(0)

# Create the output data structure based on the format
output_data = []
for i, year in enumerate(years):
    electric_component = appliance_values[i] * 0.5
    construction_component = appliance_values[i] * 0.3  
    furniture_component = appliance_values[i] * 0.2
    
    if i == 0:
        # First year has no growth rate
        row = [year, electric_component, construction_component, furniture_component, None]
    else:
        row = [year, electric_component, construction_component, furniture_component, growth_rates[i-1]]
    output_data.append(row)

# Create new workbook for growth_rate.xlsx
output_wb = Workbook()
ws = output_wb.active
ws.title = "Sheet1"

# Write headers
headers = ["Year", "Electric 50%", "Construction 30%", "Furniture 20%", "Growth Rate"]
ws.append(headers)

# Write data
for row in output_data:
    ws.append(row)

# Save the file
output_wb.save('/ssddata/mcpbench/zhaojian/mcpbench_dev/recorded_trajectories_v2/run1/claude-4-sonnet-0514/finalpool/SingleUserTurn-excel-market-research/workspace/growth_rate.xlsx')

print("Data processing completed. Created growth_rate.xlsx")
print(f"Processed {len(years)} years of data from {min(years)} to {max(years)}")
print(f"Sample appliance values for first 3 years: {appliance_values[:3]}")
print(f"Sample growth rates for first 3 years: {growth_rates[:3]}")

# Print some verification data
print("\nVerification data:")
for i in range(min(3, len(years))):
    year = years[i]
    electric = raw_data[i][5]
    furniture = raw_data[i][8]
    construction = raw_data[i][9]
    appliance = appliance_values[i]
    print(f"Year {year}: Electric={electric}, Furniture={furniture}, Construction={construction}bn, Appliance={appliance}")